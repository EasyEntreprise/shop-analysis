# Librairies
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import altair as alt
from streamlit_extras.dataframe_explorer import dataframe_explorer
from streamlit_extras.metric_cards import style_metric_cards
from prophet import Prophet
from prophet.plot import plot_plotly, plot_components_plotly
import matplotlib.pyplot as plt
import os
import time
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


# Config page
st.set_page_config(
    page_title="SHOP BUSINESS ANALYSIS", 
    layout="wide", 
    page_icon=":material/bar_chart:"
)


st.markdown("<h1 style='text-align: center; color: blue;'> SHOP SITUATION DASHBOARD </h1>", unsafe_allow_html= True)
st.markdown("<br/>", unsafe_allow_html= True)
st.markdown("<br/>", unsafe_allow_html= True)
st.markdown("<h6 style='text-align: center; color: red;'> Welcome in our shop situation dashboard. This dashboard is important for following the your purchases and sales."
"In this "
"</h6>", unsafe_allow_html= True)

st.markdown("___")

######################################
## Boutons fermeture et rederamarrage
#####################################
ferm, rederm = st.columns(2)
with ferm:
    # ------------------------------
    # 🔴 Bouton Fermer (avec confirmation)
    # ------------------------------
    if "confirm_exit" not in st.session_state:
        st.session_state.confirm_exit = False

    if st.session_state.confirm_exit:
        st.warning("❗ Are you sure you want to stop using the application altogether ?")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("✅ Yes, stop "):
                st.error("🛑 Closing the application... ")
                time.sleep(1)
                os._exit(0)
        with col2:
            if st.button("❌ No, cancel "):
                st.session_state.confirm_exit = False
    else:
        if st.button("🛑 Close application "):
            st.session_state.confirm_exit = True

with rederm:
    # ------------------------------
    # 🔄 Bouton Redémarrer (avec confirmation)
    # ------------------------------
    if "confirm_restart" not in st.session_state:
        st.session_state.confirm_restart = False

    if st.session_state.confirm_restart:
        st.warning("❗ Are you sure you want to restart the application ? ")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("✅ Yes, restart "):
                st.info("🔄 Restarting the application...")
                time.sleep(1)
                st.experimental_rerun()
        with col2:
            if st.button("❌ No, cancel"):
                st.session_state.confirm_restart = False
    else:
        if st.button("🔄 Restart application "):
            st.session_state.confirm_restart = True
st.markdown("___")

####################################
### Fonction de lecture de fichier
####################################

def read_file(file):
    """Lit automatiquement Excel ou CSV selon le type du fichier."""
    if file is None:
        return None
    
    filename = file.name.lower()

    if filename.endswith(".csv"):
        return pd.read_csv(file)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        return pd.read_excel(file)
    else:
        st.error("Format not supported. Use Excel. (.xlsx/.xls) ou CSV.")
        return None


##################
# Load dataset
#######
file = st.file_uploader("📂 Inserer votre fichier Excel en appuyant sur le bouton 'Browse files'", type=["xlsx","xls", "xlsm"])

if file is not None:
    
    #feuilles = ["Inventaires", "Entrées", "Sorties"]
    feuilles = ["Inventaires"]

    wb = load_workbook(file, data_only=True)
    winventory = wb["Inventaires"]
    wenter = wb["Entrées"]
    ws = wb["Sorties"]

    # Nom exact du tableau Excel
    tableIntory = winventory.tables["Inventaire"]
    tableEnter = wenter.tables["ListeEntree"]
    tableSorti = ws.tables["ListeSortie"]

    # Décomposer la plage du tableau
    min_col_inv, min_row_inv, max_col_inv, max_row_inv = range_boundaries(tableIntory.ref)
    min_col_enter, min_row_enter, max_col_enter, max_row_enter = range_boundaries(tableEnter.ref)
    min_col_out, min_row_out, max_col_out, max_row_out = range_boundaries(tableSorti.ref)

    #dataset_full = pd.read_excel(file, sheet_name= feuilles, usecols=table_range, engine="openpyxl")
    
    inventaire = pd.read_excel(
        file,
        sheet_name="Inventaires",
        skiprows=min_row_inv - 1,                  # sauter les lignes avant le tableau
        nrows=max_row_inv - min_row_inv + 1,       # nombre de lignes du tableau
        usecols=range(min_col_inv - 1, max_col_inv),
        engine="openpyxl"
    )

    entrer = pd.read_excel(
        file,
        sheet_name="Entrées",
        skiprows=min_row_enter - 1,                    # sauter les lignes avant le tableau
        nrows=max_row_enter - min_row_enter + 1,       # nombre de lignes du tableau
        usecols=range(min_col_enter - 1, max_col_enter),
        engine="openpyxl"
    )

    sortie = pd.read_excel(
        file,
        sheet_name="Sorties",
        skiprows=min_row_out - 1,                  # sauter les lignes avant le tableau
        nrows=max_row_out - min_row_out + 1,       # nombre de lignes du tableau
        usecols=range(min_col_out - 1, max_col_out),
        engine="openpyxl"
    )

    
    ####################
    # Creation des dates
    ####################

    # Convertir en date
    entrer["Date"] = pd.to_datetime(entrer["Date"], errors = "coerce")
    sortie["Date"] = pd.to_datetime(sortie["Date"], errors = "coerce")

    # Extraire le mois et année ensemble
    entrer["Mois"] = entrer["Date"].dt.strftime("%B %Y")  #entrer["Date"].dt.month ou entrer["Date"].dt.strftime("%B")
    sortie["Mois"] = sortie["Date"].dt.strftime("%B %Y")  #.dt.month_name(locale= "fr_FR")

    #

    # Extraire l'année
    entrer["Année"] = entrer["Date"].dt.year
    sortie["Année"] = sortie["Date"].dt.year

    #st.dataframe(inventaire)
    #st.dataframe(entrer)
    #st.dataframe(sortie)

    st.markdown("___")

    ###############################
    ## METRIC
    ###############################
    st.subheader("1. VISUALISATION NUMERIQUE", divider="rainbow")

    col1, col2, col3, col4, col5, col6 = st.columns(6)

    with col1 :

        stock_final_inv = inventaire["Stock final"].sum()
        st.metric(label= "Stock total :", value= f"{stock_final_inv} Pcs")

        st.text("Stock par catégories")

        stock_final_graph = inventaire.groupby("Catégorie", as_index= False)["Stock final"].sum()
        fig_one = go.Figure(data = [go.Pie(labels = stock_final_graph["Catégorie"], values= stock_final_graph["Stock final"], opacity=0.5)])
        fig_one.update_traces (hoverinfo='label+percent', textfont_size=15,textinfo= 'label+percent', pull= [0.05, 0, 0, 0, 0],marker_line=dict(color='#FFFFFF', width=2))
        st.plotly_chart(fig_one)
        

    with col2 :
        valeur_stock_inv = inventaire["Valeur"].sum()
        st.metric(label= "Valeur de stock :", value= f"{valeur_stock_inv} $")

        st.text("Valeur par catégories")

        valeur_final_graph = inventaire.groupby("Catégorie", as_index= False)["Valeur"].sum()
        fig_two = go.Figure(data = [go.Pie(labels = valeur_final_graph["Catégorie"], values= valeur_final_graph["Valeur"], opacity=0.5)])
        fig_two.update_traces (hoverinfo='label+percent', textfont_size=15,textinfo= 'label+percent', pull= [0.05, 0, 0, 0, 0],marker_line=dict(color='#FFFFFF', width=2))
        st.plotly_chart(fig_two)
        

    with col3 :
        total_entree = entrer["Quantité"].sum()
        st.metric(label= "Total des entrées :", value= f"{total_entree} Pcs")

        st.text("Total des entrées par catégories")

        total_entrer_graph = entrer.groupby("Catégorie", as_index= False)["Quantité"].sum()
        fig_three = go.Figure(data = [go.Pie(labels = total_entrer_graph["Catégorie"], values= total_entrer_graph["Quantité"], opacity=0.5)])
        fig_three.update_traces (hoverinfo='label+percent', textfont_size=15,textinfo= 'label+percent', pull= [0.05, 0, 0, 0, 0],marker_line=dict(color='#FFFFFF', width=2))
        st.plotly_chart(fig_three)
        

    with col4 :
        valeur_entree = entrer["Total"].sum()
        st.metric(label= "Valeur de stock :", value= f"{valeur_entree} $")

        st.text("Valeur des entrées catégories")

        valeur_entrer_graph = entrer.groupby("Catégorie", as_index= False)["Total"].sum()
        fig_for = go.Figure(data = [go.Pie(labels = valeur_entrer_graph["Catégorie"], values= valeur_entrer_graph["Total"], opacity=0.5)])
        fig_for.update_traces (hoverinfo='label+percent', textfont_size=15,textinfo= 'label+percent', pull= [0.05, 0, 0, 0, 0],marker_line=dict(color='#FFFFFF', width=2))
        st.plotly_chart(fig_for)
        

    with col5 :
        total_sortie = sortie["Quantité"].sum()
        st.metric(label= "Total des sorties :", value= f"{total_sortie} Pcs")

        st.text("Total des sorties par catégories")

        total_sortie_graph = sortie.groupby("Catégorie", as_index= False)["Quantité"].sum()
        fig_five = go.Figure(data = [go.Pie(labels = total_sortie_graph["Catégorie"], values= total_sortie_graph["Quantité"], opacity=0.5)])
        fig_five.update_traces (hoverinfo='label+percent', textfont_size=15,textinfo= 'label+percent', pull= [0.05, 0, 0, 0, 0],marker_line=dict(color='#FFFFFF', width=2))
        st.plotly_chart(fig_five)
        

    with col6 :
        valeur_sortie = sortie["Total"].sum()
        st.metric(label= "Valeur de stock :", value= f"{valeur_sortie} $")

        st.text("Valeur des entrées catégories")

        valeur_sortie_graph = sortie.groupby("Catégorie", as_index= False)["Total"].sum()
        fig_six = go.Figure(data = [go.Pie(labels = valeur_sortie_graph["Catégorie"], values= valeur_sortie_graph["Total"], opacity=0.5)])
        fig_six.update_traces (hoverinfo='label+percent', textfont_size=15,textinfo= 'label+percent', pull= [0.05, 0, 0, 0, 0],marker_line=dict(color='#FFFFFF', width=2))
        st.plotly_chart(fig_six)
        

    # Style the metric
    style_metric_cards(background_color="#636363", border_left_color="#a8ff78", border_color="#9FC1FF")


    ################
    ## Graphique
    #############

    st.subheader("2. GRAPHIQUE", divider="rainbow")

    col7, col8 = st.columns(2)
    
    with col7:
        st.text("Select Date Range")
        start_date = st.date_input(label="Start Dates")

    with col8:
        st.text("Select Date Range")
        en_date = st.date_input(label="End Dates")

    # Provide a message for selected date range 
    st.success(" you have choosen analytics from: "+str(start_date)+" to "+str(en_date))

    # Filtre dates
    date_entrer = entrer[(entrer["Date"]>=str(start_date)) & (entrer["Date"]<=str(en_date))]
    date_sortie = sortie[(sortie["Date"]>=str(start_date)) & (sortie["Date"]<=str(en_date))]

    ##############
    # 2.1. Yearly

    st.header("2.1. Yearly", divider="gray")

    col9, col10 = st.columns(2)

    with col9 :
        yearly_enter = entrer.groupby("Année", as_index= False)["Quantité"].sum()

        fig_seven = px.line(yearly_enter, x="Année", y="Quantité", text= "Quantité", title="La valeur totale des achats par Ans (Pcs)")
        fig_seven.update_traces(textposition = 'top center')
        st.plotly_chart(fig_seven)

        st.markdown("___")

        yearly_enter_valeur = entrer.groupby("Année", as_index= False)["Total"].sum()

        fig_nine = px.line(yearly_enter_valeur, x="Année", y="Total", text="Total", title="La valeur totale des achats par Ans ($)")
        fig_nine.update_traces(textposition = 'top center')
        st.plotly_chart(fig_nine)
        

    with col10 :
         yearly_sortie = sortie.groupby("Année", as_index= False)["Quantité"].sum()
         
         fig_eigth = px.line(yearly_sortie, x="Année", y="Quantité", text="Quantité", title="La valeur totale des ventes par Ans (Pcs)")
         fig_eigth.update_traces(textposition = 'top center')
         st.plotly_chart(fig_eigth)

         st.markdown("___")

         yearly_sortie_valeur = sortie.groupby("Année", as_index= False)["Total"].sum()

         fig_ten = px.line(yearly_sortie_valeur, x="Année", y="Total", text="Total", title="La valeur totale des ventes par Ans ($)")
         fig_ten.update_traces(textposition = 'top center')
         st.plotly_chart(fig_ten)

    ##############
    # 2.1. Yearly
    st.header("2.2. Monthly", divider="gray")

    col11, col12 = st.columns(2)

    with col11 :
        monthly_enter = date_entrer.groupby("Mois", as_index= False)["Quantité"].sum()

        fig_eleven = px.line(monthly_enter, x="Mois", y="Quantité", text="Quantité", title="La valeur totale des achats par Mois (Pcs)")
        fig_eleven.update_traces(textposition = 'top center')
        st.plotly_chart(fig_eleven)

        st.markdown("___")

        monthly_enter_valeur = date_entrer.groupby("Mois", as_index= False)["Total"].sum()

        fig_twelve = px.line(monthly_enter_valeur, x="Mois", y="Total", text="Total", title="La valeur totale des achats par Mois ($)")
        fig_twelve.update_traces(textposition = 'top center')
        st.plotly_chart(fig_twelve)
        

    with col12 :
         monthly_sortie = date_sortie.groupby("Mois", as_index= False)["Quantité"].sum()
         
         fig_thirdteen = px.line(monthly_sortie, x="Mois", y="Quantité", text="Quantité", title="La valeur totale des ventes par Mois (Pcs)")
         fig_thirdteen.update_traces(textposition = 'top center')
         st.plotly_chart(fig_thirdteen)

         st.markdown("___")

         monthly_sortie_valeur = date_sortie.groupby("Mois", as_index= False)["Total"].sum()

         fig_forteen = px.line(monthly_sortie_valeur, x="Mois", y="Total", text="Total", title="La valeur totale des ventes par Mois ($)")
         fig_forteen.update_traces(textposition = 'top center')
         st.plotly_chart(fig_forteen)


    ##############
    # 2.3. Others

    st.subheader("2.3. OTHERS", divider="gray")

    col13, col14, col15 = st.columns(3)

    with col13 :
        # Marchandises à approvisionner
        st.text("2.3.1. Marchandises à approvisionner")

        statue_autoriser = ["Stock faible", "Produit non disponible"]
        statue = inventaire[inventaire["Statue"].isin(statue_autoriser)]
        groupby_statue = statue.groupby(["Référence", "Statue"]).size().reset_index(name="nombre")

        fig_zero = px.bar(groupby_statue, x="Référence", y="nombre", color="Statue")
        fig_zero.update_traces(textposition = 'outside')
        st.plotly_chart(fig_zero)


        # Les modeles avec plus des ventes
        top_modeles = (date_sortie.groupby("Référence")["Quantité"].sum().reset_index().sort_values(by="Quantité", ascending= False))
       
        # Afficher les 10 modeles le plus vendus
        st.text("2.3.1.1 Top 10 des modèles les plus vendus")
        st.dataframe(top_modeles.head(10))

        st.bar_chart(
            top_modeles.head(10).set_index("Référence")
        )


    with col14 :
        # Achats par catégories avec selection
        st.text("2.3.2. Achats par catégories")

        category_entrer = date_entrer["Catégorie"].unique()
        all_entrer = ["Toutes catégories"] + sorted(date_entrer["Catégorie"].dropna().unique().tolist())
        
        selector = st.selectbox("Choose your Category for Purchasing :", all_entrer)

        if selector == "Toutes catégories" :
            ach_selector = date_entrer

        else:
            ach_selector = date_entrer[date_entrer["Catégorie"] == selector]

        groupe_by_achat = ach_selector.groupby("Référence", as_index= False)["Quantité"].sum()

        fig_fifteen = px.bar(groupe_by_achat, x="Référence", y="Quantité", text="Quantité", color="Référence")
        fig_fifteen.update_traces(textposition = 'outside')
        st.plotly_chart(fig_fifteen)

        reference_entrer = date_entrer["Référence"].unique()
        selector_ref = st.selectbox("Choose your model for Purchasing :", reference_entrer)
        ach_ref = date_entrer[date_entrer["Référence"] == selector_ref]
        
        groupe_by_achat = ach_ref.groupby(["Référence", "Mois"], as_index= False)["Quantité"].sum()
        fig_sixteen = px.line(groupe_by_achat, x="Mois", y="Quantité", text="Quantité")
        fig_sixteen.update_traces(textposition = 'top center')
        st.plotly_chart(fig_sixteen)
    
    with col15 :
        # Sortie par catégories avec selection
        st.text("2.3.3. Ventes par catégories")

        category_sortie = date_sortie["Catégorie"].unique()
        all_sortie = ["Toutes catégories"] + sorted(date_sortie["Catégorie"].dropna().unique().tolist())
        
        selector = st.selectbox("Choose your Category for saling :", all_sortie)

        if selector == "Toutes catégories" :
            sortie_selector = date_sortie

        else:
            sortie_selector = date_sortie[date_sortie["Catégorie"] == selector]
        
        groupe_by_vente = sortie_selector.groupby("Référence", as_index= False)["Quantité"].sum()

        fig_seventeen = px.bar(groupe_by_vente, x="Référence", y="Quantité", text="Quantité", color="Référence")
        fig_seventeen.update_traces(textposition = 'outside')
        st.plotly_chart(fig_seventeen)

        reference_sortie = date_sortie["Référence"].unique()
        selector_ref_s = st.selectbox("Choose your model for Saling :", reference_sortie)
        vente_ref = date_sortie[date_sortie["Référence"] == selector_ref_s]
        
        groupe_by_vente = vente_ref.groupby(["Référence", "Mois"], as_index= False)["Quantité"].sum()
        fig_eighteen = px.line(groupe_by_vente, x="Mois", y="Quantité", text="Quantité")
        fig_eighteen.update_traces(textposition = 'top center')
        st.plotly_chart(fig_eighteen)
        